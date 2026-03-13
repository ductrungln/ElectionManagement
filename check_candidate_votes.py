import openpyxl
wb = openpyxl.load_workbook('MauNhap_Excel_tinh_7_bau_4.xlsx')
ws = wb.active

print('=== Candidate Votes Analysis (Row 105 & others) ===\n')

print('Row 105 values:')
for col in range(1, 12):
    col_letter = openpyxl.utils.get_column_letter(col)
    val = ws[f'{col_letter}105'].value
    print(f'{col_letter}105: {val}')

print(f'\n=== Testing Formulas ===')
print(f'Candidate 1: C105 - D105 = {ws["C105"].value} - {ws["D105"].value}')
print(f'Candidate 2: C105 - E105 = {ws["C105"].value} - {ws["E105"].value}')
print(f'Candidate 3: C105 - F105 = {ws["C105"].value} - {ws["F105"].value}')
print(f'Candidate 4: C105 - G105 = {ws["C105"].value} - {ws["G105"].value}')
print(f'Candidate 5: C105 - H105 = {ws["C105"].value} - {ws["H105"].value}')
print(f'Candidate 6: C105 - I105 = {ws["C105"].value} - {ws["I105"].value}')
print(f'Candidate 7: C105 - J105 = {ws["C105"].value} - {ws["J105"].value}')

print(f'\n=== Check row 104 as well ===')
for col in range(1, 12):
    col_letter = openpyxl.utils.get_column_letter(col)
    val = ws[f'{col_letter}104'].value
    print(f'{col_letter}104: {val}')
